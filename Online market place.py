
import sqlite3
import pandas as pd
from openpyxl import Workbook, load_workbook
from datetime import datetime
import hashlib

# Global variables
PRODUCTS = {
    "Pen": {"price": 15, "quantity": 100},
    "Notebook": {"price": 30, "quantity": 50},
    "Pencil": {"price": 10, "quantity": 75},
    "setsqure": {"price": 45, "quantity": 75},
    "compass": {"price": 100, "quantity": 50},
    "marker": {"price": 25, "quantity": 200},
    "color": {"price": 100, "quantity": 150}
}

# Global dictionary to store usernames and hashed passwords
USERS = {
    "admin": "5f4dcc3b5aa765d61d8327deb882cf99",  # "password" hashed using MD5
    "user": "7c4a8d09ca3762af61e59520943dc26494f8941b"  # "123456" hashed using SHA-1
}

# Global dictionary to store failed login attempts
FAILED_LOGIN_ATTEMPTS = {}

# Maximum number of allowed failed login attempts before lockout
MAX_FAILED_ATTEMPTS = 3

# Function to hash passwords
def hash_password(password):
    # You can choose a stronger hashing algorithm like SHA-256 or bcrypt for better security
    return hashlib.md5(password.encode()).hexdigest()

# Function to create SQLite database
def create_database():
    conn = sqlite3.connect('store.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS sheet
                 (name TEXT, mobile TEXT, product TEXT, price REAL)''')
    conn.commit()
    conn.close()
    print("Database created successfully.")

# Function to insert sales data into SQLite database
def insert_sales(customer_name, contact, product, total_price):
    conn = sqlite3.connect('store.db')
    c = conn.cursor()
    c.execute("INSERT INTO sheet (name, mobile, product, price) VALUES (?, ?, ?, ?)", (customer_name, contact, product, total_price))
    conn.commit()
    conn.close()

# Function to create an Excel sheet to store sales data
def create_sales_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Timestamp", "Customer Name", "Contact", "Product", "Quantity", "Unit Price", "Total Price", "Discount", "Final Price"])
    wb.save("sales.xlsx")
    print("Sales Excel sheet created successfully.")

# Function to update sales data in the Excel sheet
def update_sales(customer_name, contact, product, quantity, unit_price, total_price, discount, final_price):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = load_workbook("sales.xlsx")
    ws = wb.active
    ws.append([timestamp, customer_name, contact, product, quantity, unit_price, total_price, discount, final_price])
    wb.save("sales.xlsx")

# Function to create an Excel sheet to store product information
def create_product_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Inventory"
    ws.append(["Product", "Price", "Quantity"])
    for product, info in PRODUCTS.items():
        ws.append([product, info['price'], info['quantity']])
    wb.save("products.xlsx")
    print("Product inventory Excel sheet created successfully.")

# Function to update product inventory
def update_inventory(product, quantity_sold):
    PRODUCTS[product]["quantity"] -= quantity_sold
    wb = load_workbook("products.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == product:
            row[2].value = PRODUCTS[product]["quantity"]
            break
    wb.save("products.xlsx")

# Function to calculate discount based on total price
def calculate_discount(total_price):
    if total_price >= 100:
        return 0.1  # 10% discount for total price >= $100
    elif total_price >= 50:
        return 0.05  # 5% discount for total price >= $50
    else:
        return 0

# Function to display available products
def display_products():
    print("Available Products:")
    for product, info in PRODUCTS.items():
        print(f"{product}: ${info['price']} (Quantity: {info['quantity']})")

# Function to sell products to customers
def sell_product():
    print("Welcome to the Stationary Store!")
    customer_name = input("Enter customer name: ")
    contact = input("Enter customer contact: ")
    
    while True:
        display_products()
        product = input("Enter product name (or type 'done' to finish shopping): ")
        if product.lower() == 'done':
            break
        if product not in PRODUCTS:
            print("Invalid product.")
            continue
        quantity = int(input("Enter quantity: "))
        if quantity > PRODUCTS[product]["quantity"]:
            print("Insufficient stock.")
            continue
        
        unit_price = PRODUCTS[product]["price"]
        total_price = quantity * unit_price
        discount = calculate_discount(total_price)
        final_price = total_price - (total_price * discount)
        
        print(f"Total Price: ${total_price}")
        print(f"Discount applied: {discount * 100}%")
        print(f"Final Price after discount: ${final_price}")
        
        insert_sales(customer_name, contact, product, final_price)
        update_sales(customer_name, contact, product, quantity, unit_price, total_price, discount, final_price)
        update_inventory(product, quantity)
        print("Product added to cart.")

    print("Sales data, database, and inventory updated successfully.")

# Function for user login
def login():
    print("Login or Create Account")
    while True:
        print("1. Login")
        print("2. Create Account")
        print("3. Exit")
        choice = input("Enter your choice: ")

        if choice == '1':
            username = input("Enter username: ")
            # Check if the username exists
            if username not in USERS:
                print("Invalid username.")
                continue
            # Check if the account is locked due to too many failed attempts
            if username in FAILED_LOGIN_ATTEMPTS and FAILED_LOGIN_ATTEMPTS[username] >= MAX_FAILED_ATTEMPTS:
                print("Account locked. Too many failed login attempts.")
                continue
            password = input("Enter password: ")
            hashed_password = hash_password(password)
            # Check if the password matches the hashed password
            if USERS[username] == hashed_password:
                print("Login successful.")
                # Reset failed login attempts for the user
                FAILED_LOGIN_ATTEMPTS.pop(username, None)
                return True
            else:
                print("Incorrect password.")
                # Increment failed login attempts count for the user
                FAILED_LOGIN_ATTEMPTS[username] = FAILED_LOGIN_ATTEMPTS.get(username, 0) + 1
                continue
        elif choice == '2':
            username = input("Enter a new username: ")
            if username in USERS:
                print("Username already exists. Please choose another username.")
                continue
            password = input("Enter a password: ")
            hashed_password = hash_password(password)
            USERS[username] = hashed_password
            print("Account created successfully.")
            # Proceed with login after creating account
            return True
        elif choice == '3':
            print("Exiting...")
            return False
        else:
            print("Invalid choice. Please enter a valid option.")

            continue
def display_menu():
    print("Welcome to the Stationary Store!")
    print("1. Display available products")
    print("2. Purchase products")
    print("3. Exit")

# Main function
def main():
    if login():
        create_database()
        create_sales_excel()
        create_product_excel()

        while True:
            display_menu()
            choice = input("Enter your choice: ")

            if choice == '1':
                display_products()
            elif choice == '2':
                sell_product()
            elif choice == '3':
                print("Thank you for visiting. Have a great day!")
                break
            else:
                print("Invalid choice. Please enter a valid option.")

if __name__ == "__main__":
    main()

# Function to display menu
